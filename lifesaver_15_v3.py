import os
import re
import autoit
import time
import csv
from datetime import datetime
from easygui import msgbox
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
import statistics


def make_pack_folder(p_number, w_speed, w_style):
    root = "C:\\Users\\Eric\\Desktop\\15sec_Capture\\LP Power Data\\"
    target = root + "LP-{0}".format(p_number)
    if not os.path.exists(target):
        print(" ")
        print("made new pack folder at: ", target)
        os.mkdir(target)
    return make_speed_folder(target, w_speed, w_style)


def make_speed_folder(s_target, w_speed, w_style):
    speed_target = os.path.join(s_target, "{0}MPH-{1} Walk".format(w_speed, w_style))
    if not os.path.exists(speed_target):
        os.mkdir(speed_target)
    else:
        return


def convert_csv_to_xlsx(csv_filename, dest_filename):
    wb = load_workbook("C:\\Users\\Eric\\Desktop\\15sec_Capture\\csv_transform_template.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")

    f = open(csv_filename, 'r')
    reader = csv.reader(f)
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            s = cell
            try:
                s = float(s)
            except ValueError:
                pass
            ws.cell('%s%s' % (column_letter, (row_index + 1))).value = s
    wb.save(filename=dest_filename)


def take_abs_val(sheet, column_number):
    hr = sheet.get_highest_row()
    for i in range(4, hr):
        sheet.cell(row=i, column=column_number).value = abs(sheet.cell(row=i, column=column_number).value)
    col_name = sheet.cell(row=1, column=column_number).value
    sheet.cell(row=1, column=column_number).value = "abs({0})".format(col_name)


def fix_xlsx(path):
    wb = load_workbook(filename=path)
    sheet1 = wb.get_active_sheet()
    take_abs_val(sheet1, 2)
    hr = sheet1.get_highest_row()
    pwr_list = []
    print("highest row=", hr)
    for i in range(4, hr+1):
        pwr_list.append(sheet1.cell(row=i, column=2).value)
    avg_pwr = round(statistics.mean(pwr_list), 2)
    sheet1['C1'].value = "Avg. Power"
    sheet1['C2'].value = "(W)"
    #sheet1['C4'].value = "=ROUND(AVERAGE(B4:B15050),2)"
    sheet1['C4'].value = avg_pwr
    wb.save(filename=path)


def patch_ref_database(fixed_xlsx, final_ref_doc, pack_number, walk_speed):
    wb_fixed = load_workbook(filename=fixed_xlsx)
    sheet1_fixed = wb_fixed.get_sheet_by_name("Sheet1")
    magic_num = sheet1_fixed['C4'].value
    #print("magic_num= ", magic_num)
    mod_row = 4+int(str(pack_number).lstrip("0"))
    #print("mod row= ", mod_row)

    if float(walk_speed) == 2.5:
        wb_ref = load_workbook(filename=final_ref_doc)
        packsheet_ref = wb_ref.get_sheet_by_name("packs")
        packsheet_ref['Q{}'.format(mod_row)].value = magic_num
        wb_ref.save(filename=final_ref_doc)
    if float(walk_speed) == 3.5:
        wb_ref = load_workbook(filename=final_ref_doc)
        packsheet_ref = wb_ref.get_sheet_by_name("packs")
        packsheet_ref['R{}'.format(mod_row)].value = magic_num
        wb_ref.save(filename=final_ref_doc)
    print("")
    print("This run's power was {} Watts".format(magic_num))


def main():
    for i in range(0, 2):
        print(".")
    print("Please enter the info below for the pack to be tested: ")
    print(" ")
    pack_number = ""
    while re.search(r'\d\d\d', pack_number) is None:
        pack_number = input("Pack Number (3 digits please): ")
    print(" ")
    walk_speed = ""
    while re.search(r'\d[.]\d', walk_speed) is None:
        walk_speed = input("Test Speed (ex. 2.5/3.5): ")
    print(" ")
    walk_style = ""
    while walk_style != "LP":
        if walk_style == "LP":
            break
        if walk_style == "Normal":
            break
        else:
            walk_style = input("Walk Style (LP/Normal): ")
    make_pack_folder(pack_number, walk_speed, walk_style)

    print('.....')
    print("..........Launching.....")

    autoit.auto_it_set_option('MouseCoordMode', 0)
    autoit.auto_it_set_option('SendKeyDelay', 10)
    autoit.run("C:\\Program Files (x86)\\Pico Technology\\PicoScope6\\PicoScope.exe")
    autoit.win_wait('PicoScope 6')
    autoit.win_activate('PicoScope 6')
    autoit.send('{ALT}{f}{o}')
    time.sleep(2)

    open_text = "C:\\Users\\Eric\\Desktop\\15sec_Capture\\electrical_power.psdata"
    autoit.clip_put(open_text)
    autoit.win_activate('Open')
    autoit.win_wait("Open")
    autoit.control_send('Open', "[Class:Edit;INSTANCE:1]", '^V', 0)
    autoit.send('{ENTER}')
    autoit.win_wait('PicoScope 6 - [electrical_power.psdata]')
    autoit.win_activate('PicoScope 6 - [electrical_power.psdata]')
    time.sleep(0.5)
    autoit.send('{ALT}{TAB}{TAB}{TAB}{TAB}{ENTER}{END}{UP}{ENTER}')
    autoit.win_wait('Macro Recorder')
    autoit.control_click('Macro Recorder', "[Name:_buttonImport]")
    time.sleep(0.5)

    macro_text = "C:\\Users\\Eric\\Desktop\\15sec_Capture\\15sec_pico_record_macro.psmacro"
    autoit.clip_put(macro_text)
    autoit.control_send('Open', "[Class:Edit;INSTANCE:1]", '^V', 0)
    autoit.send('{ENTER}')
    msgbox("Ready to go?")
    autoit.win_activate('Macro Recorder')
    autoit.win_wait('Macro Recorder')
    autoit.control_click('Macro Recorder', "[Name:_buttonExecute]")
    time.sleep(7.5)
    autoit.win_activate('Macro Recorder')
    autoit.send('{ESCAPE}')

    #msgbox("Stop Runnin'-- Enter power numbers into spreadsheet. Then return to this box and hit OK to continue.")
    autoit.win_activate('PicoScope 6')
    autoit.send('{ALT}{f}{a}')
    time.sleep(.5)
    i = datetime.now()
    save_text = "{}".format(os.path.join("C:\\Users\\Eric\\Desktop\\15sec_Capture\\LP Power Data\\",
                             "LP-{0}".format(pack_number), "{0}MPH-{1} Walk".format(walk_speed, walk_style),
                             "{0}MPH-{1} Walk_".format(walk_speed, walk_style) + i.strftime('%Y-%m-%d %Hh%Mm%Ss')+".psdata"))
    autoit.clip_put(save_text)
    autoit.control_send('Save As', "[Class:Edit;INSTANCE:1]", '^V', 0)
    time.sleep(0.5)
    autoit.send('{ENTER}')
    time.sleep(0.5)
    #autoit.win_activate('PicoScope 6')
    autoit.win_wait("PicoScope 6")
    autoit.mouse_click("primary", 433, 67, 2, 1)
    #autoit.control_click('PicoScope 6', "[CLASS:WindowsForms10.BUTTON.app.0.2f5a4f0_r13_ad1;INSTANCE:38]")

    autoit.clip_put(save_text)
    autoit.win_activate('PicoScope 6')
    autoit.control_click('PicoScope 6', "[CLASS:WindowsForms10.Window.8.app.0.2f5a4f0_r13_ad1;INSTANCE:55]")
    autoit.send('{ALT}{f}{a}')
    autoit.win_wait("Save As")
    #autoit.send("{^a}{DELETE}", 1)
    autoit.control_send('Save As', "[Class:Edit;INSTANCE:1]", '^V', 0)
    autoit.control_click('Save As', "[CLASS:ComboBox;INSTANCE:3]")
    autoit.send("{HOME}{DOWN}{DOWN}{ENTER}")
    autoit.control_click("Save As", "[NAME:_currentBufferRadio]")
    time.sleep(0.5)
    autoit.send('{ENTER}')
    time.sleep(5)
    autoit.process_close('PicoScope.exe')

    csv_path = save_text.replace(".psdata", ".csv")
    dest_filename = save_text.replace(".psdata", ".xlsx")
    convert_csv_to_xlsx(csv_path, dest_filename)
    fix_xlsx(dest_filename)
    ref_stopgap_spreadsheet = 'C:\\Users\\Eric\\Desktop\\15sec_Capture\\REF_EMOLLE_Database_StopGap.xlsx'
    patch_ref_database(dest_filename, ref_stopgap_spreadsheet, pack_number, walk_speed)

    print("........................... Testing Complete.")


if __name__ == '__main__':
    main()
